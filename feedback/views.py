from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from .models import Lesson, FeedbackRecord, GuideBook, QuizSet, Question, StudentAnswer
import gspread
from google.oauth2.service_account import Credentials
from datetime import datetime
import os
import json
import threading
import re
import io
import base64
from collections import OrderedDict

def get_sheet():
    scope = [
        'https://spreadsheets.google.com/feeds',
        'https://www.googleapis.com/auth/drive'
    ]
    creds_json = os.environ.get('GOOGLE_CREDENTIALS')
    if creds_json:
        creds_dict = json.loads(creds_json)
        creds = Credentials.from_service_account_info(creds_dict, scopes=scope)
    else:
        creds_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'credentials.json')
        creds = Credentials.from_service_account_file(creds_path, scopes=scope)
    client = gspread.authorize(creds)
    sheet = client.open_by_key('1A7awsXWOu-WPiRjY6vk8rPEhBh8PpGiTp3pGAlellsM').sheet1
    return sheet

LESSONS = [
    # 8반 (수목금)
    {'id': 101, 'title': '8반 - 1차시', 'date': '2026-04-15', 'class': '8반'},
    {'id': 102, 'title': '8반 - 2차시', 'date': '2026-04-16', 'class': '8반'},
    {'id': 103, 'title': '8반 - 3차시', 'date': '2026-04-17', 'class': '8반'},
    {'id': 104, 'title': '8반 - 4차시', 'date': '2026-04-22', 'class': '8반'},
    {'id': 105, 'title': '8반 - 5차시', 'date': '2026-04-23', 'class': '8반'},
    {'id': 106, 'title': '8반 - 6차시', 'date': '2026-04-24', 'class': '8반'},
    {'id': 107, 'title': '8반 - 7차시', 'date': '2026-04-29', 'class': '8반'},
    {'id': 108, 'title': '8반 - 8차시', 'date': '2026-04-30', 'class': '8반'},
    {'id': 109, 'title': '8반 - 9차시', 'date': '2026-05-01', 'class': '8반'},
    {'id': 110, 'title': '8반 - 10차시', 'date': '2026-05-07', 'class': '8반'},
    {'id': 111, 'title': '8반 - 11차시', 'date': '2026-05-08', 'class': '8반'},
    {'id': 112, 'title': '8반 - 12차시', 'date': '2026-05-09', 'class': '8반'},
    {'id': 113, 'title': '8반 - 13차시', 'date': '2026-05-13', 'class': '8반'},
    {'id': 114, 'title': '8반 - 14차시', 'date': '2026-05-14', 'class': '8반'},
    {'id': 115, 'title': '8반 - 15차시', 'date': '2026-05-15', 'class': '8반'},
    {'id': 116, 'title': '8반 - 16차시', 'date': '2026-05-20', 'class': '8반'},
    {'id': 117, 'title': '8반 - 17차시', 'date': '2026-05-21', 'class': '8반'},
    {'id': 118, 'title': '8반 - 18차시', 'date': '2026-05-22', 'class': '8반'},
    {'id': 119, 'title': '8반 - 19차시', 'date': '2026-05-27', 'class': '8반'},
    {'id': 120, 'title': '8반 - 20차시', 'date': '2026-05-28', 'class': '8반'},
    {'id': 121, 'title': '8반 - 21차시', 'date': '2026-05-29', 'class': '8반'},
    {'id': 122, 'title': '8반 - 22차시', 'date': '2026-06-03', 'class': '8반'},
    {'id': 123, 'title': '8반 - 23차시', 'date': '2026-06-04', 'class': '8반'},
    {'id': 124, 'title': '8반 - 24차시', 'date': '2026-06-05', 'class': '8반'},
    {'id': 125, 'title': '8반 - 25차시', 'date': '2026-06-10', 'class': '8반'},
    {'id': 126, 'title': '8반 - 26차시', 'date': '2026-06-11', 'class': '8반'},
    {'id': 127, 'title': '8반 - 27차시', 'date': '2026-06-12', 'class': '8반'},
    {'id': 128, 'title': '8반 - 28차시', 'date': '2026-06-17', 'class': '8반'},
    {'id': 129, 'title': '8반 - 29차시', 'date': '2026-06-18', 'class': '8반'},
    {'id': 130, 'title': '8반 - 30차시', 'date': '2026-06-19', 'class': '8반'},
    {'id': 131, 'title': '8반 - 31차시', 'date': '2026-06-24', 'class': '8반'},
    {'id': 132, 'title': '8반 - 32차시', 'date': '2026-06-25', 'class': '8반'},
    {'id': 133, 'title': '8반 - 33차시', 'date': '2026-06-26', 'class': '8반'},
    {'id': 134, 'title': '8반 - 34차시', 'date': '2026-07-01', 'class': '8반'},
    {'id': 135, 'title': '8반 - 35차시', 'date': '2026-07-02', 'class': '8반'},
    {'id': 136, 'title': '8반 - 36차시', 'date': '2026-07-03', 'class': '8반'},
    {'id': 137, 'title': '8반 - 37차시', 'date': '2026-07-08', 'class': '8반'},
    {'id': 138, 'title': '8반 - 38차시', 'date': '2026-07-09', 'class': '8반'},
    {'id': 139, 'title': '8반 - 39차시', 'date': '2026-07-10', 'class': '8반'},
    # 9반 (수목금)
    {'id': 201, 'title': '9반 - 1차시', 'date': '2026-04-15', 'class': '9반'},
    {'id': 202, 'title': '9반 - 2차시', 'date': '2026-04-16', 'class': '9반'},
    {'id': 203, 'title': '9반 - 3차시', 'date': '2026-04-17', 'class': '9반'},
    {'id': 204, 'title': '9반 - 4차시', 'date': '2026-04-22', 'class': '9반'},
    {'id': 205, 'title': '9반 - 5차시', 'date': '2026-04-23', 'class': '9반'},
    {'id': 206, 'title': '9반 - 6차시', 'date': '2026-04-24', 'class': '9반'},
    {'id': 207, 'title': '9반 - 7차시', 'date': '2026-04-29', 'class': '9반'},
    {'id': 208, 'title': '9반 - 8차시', 'date': '2026-04-30', 'class': '9반'},
    {'id': 209, 'title': '9반 - 9차시', 'date': '2026-05-01', 'class': '9반'},
    {'id': 210, 'title': '9반 - 10차시', 'date': '2026-05-07', 'class': '9반'},
    {'id': 211, 'title': '9반 - 11차시', 'date': '2026-05-08', 'class': '9반'},
    {'id': 212, 'title': '9반 - 12차시', 'date': '2026-05-09', 'class': '9반'},
    {'id': 213, 'title': '9반 - 13차시', 'date': '2026-05-13', 'class': '9반'},
    {'id': 214, 'title': '9반 - 14차시', 'date': '2026-05-14', 'class': '9반'},
    {'id': 215, 'title': '9반 - 15차시', 'date': '2026-05-15', 'class': '9반'},
    {'id': 216, 'title': '9반 - 16차시', 'date': '2026-05-20', 'class': '9반'},
    {'id': 217, 'title': '9반 - 17차시', 'date': '2026-05-21', 'class': '9반'},
    {'id': 218, 'title': '9반 - 18차시', 'date': '2026-05-22', 'class': '9반'},
    {'id': 219, 'title': '9반 - 19차시', 'date': '2026-05-27', 'class': '9반'},
    {'id': 220, 'title': '9반 - 20차시', 'date': '2026-05-28', 'class': '9반'},
    {'id': 221, 'title': '9반 - 21차시', 'date': '2026-05-29', 'class': '9반'},
    {'id': 222, 'title': '9반 - 22차시', 'date': '2026-06-03', 'class': '9반'},
    {'id': 223, 'title': '9반 - 23차시', 'date': '2026-06-04', 'class': '9반'},
    {'id': 224, 'title': '9반 - 24차시', 'date': '2026-06-05', 'class': '9반'},
    {'id': 225, 'title': '9반 - 25차시', 'date': '2026-06-10', 'class': '9반'},
    {'id': 226, 'title': '9반 - 26차시', 'date': '2026-06-11', 'class': '9반'},
    {'id': 227, 'title': '9반 - 27차시', 'date': '2026-06-12', 'class': '9반'},
    {'id': 228, 'title': '9반 - 28차시', 'date': '2026-06-17', 'class': '9반'},
    {'id': 229, 'title': '9반 - 29차시', 'date': '2026-06-18', 'class': '9반'},
    {'id': 230, 'title': '9반 - 30차시', 'date': '2026-06-19', 'class': '9반'},
    {'id': 231, 'title': '9반 - 31차시', 'date': '2026-06-24', 'class': '9반'},
    {'id': 232, 'title': '9반 - 32차시', 'date': '2026-06-25', 'class': '9반'},
    {'id': 233, 'title': '9반 - 33차시', 'date': '2026-06-26', 'class': '9반'},
    {'id': 234, 'title': '9반 - 34차시', 'date': '2026-07-01', 'class': '9반'},
    {'id': 235, 'title': '9반 - 35차시', 'date': '2026-07-02', 'class': '9반'},
    {'id': 236, 'title': '9반 - 36차시', 'date': '2026-07-03', 'class': '9반'},
    {'id': 237, 'title': '9반 - 37차시', 'date': '2026-07-08', 'class': '9반'},
    {'id': 238, 'title': '9반 - 38차시', 'date': '2026-07-09', 'class': '9반'},
    {'id': 239, 'title': '9반 - 39차시', 'date': '2026-07-10', 'class': '9반'},
    # 10반 (월화목)
    {'id': 301, 'title': '10반 - 1차시', 'date': '2026-04-13', 'class': '10반'},
    {'id': 302, 'title': '10반 - 2차시', 'date': '2026-04-14', 'class': '10반'},
    {'id': 303, 'title': '10반 - 3차시', 'date': '2026-04-16', 'class': '10반'},
    {'id': 304, 'title': '10반 - 4차시', 'date': '2026-04-20', 'class': '10반'},
    {'id': 305, 'title': '10반 - 5차시', 'date': '2026-04-21', 'class': '10반'},
    {'id': 306, 'title': '10반 - 6차시', 'date': '2026-04-23', 'class': '10반'},
    {'id': 307, 'title': '10반 - 7차시', 'date': '2026-04-27', 'class': '10반'},
    {'id': 308, 'title': '10반 - 8차시', 'date': '2026-04-28', 'class': '10반'},
    {'id': 309, 'title': '10반 - 9차시', 'date': '2026-04-30', 'class': '10반'},
    {'id': 310, 'title': '10반 - 10차시', 'date': '2026-05-04', 'class': '10반'},
    {'id': 311, 'title': '10반 - 11차시', 'date': '2026-05-06', 'class': '10반'},
    {'id': 312, 'title': '10반 - 12차시', 'date': '2026-05-07', 'class': '10반'},
    {'id': 313, 'title': '10반 - 13차시', 'date': '2026-05-11', 'class': '10반'},
    {'id': 314, 'title': '10반 - 14차시', 'date': '2026-05-12', 'class': '10반'},
    {'id': 315, 'title': '10반 - 15차시', 'date': '2026-05-14', 'class': '10반'},
    {'id': 316, 'title': '10반 - 16차시', 'date': '2026-05-18', 'class': '10반'},
    {'id': 317, 'title': '10반 - 17차시', 'date': '2026-05-19', 'class': '10반'},
    {'id': 318, 'title': '10반 - 18차시', 'date': '2026-05-21', 'class': '10반'},
    {'id': 319, 'title': '10반 - 19차시', 'date': '2026-05-25', 'class': '10반'},
    {'id': 320, 'title': '10반 - 20차시', 'date': '2026-05-26', 'class': '10반'},
    {'id': 321, 'title': '10반 - 21차시', 'date': '2026-05-28', 'class': '10반'},
    {'id': 322, 'title': '10반 - 22차시', 'date': '2026-06-01', 'class': '10반'},
    {'id': 323, 'title': '10반 - 23차시', 'date': '2026-06-02', 'class': '10반'},
    {'id': 324, 'title': '10반 - 24차시', 'date': '2026-06-04', 'class': '10반'},
    {'id': 325, 'title': '10반 - 25차시', 'date': '2026-06-08', 'class': '10반'},
    {'id': 326, 'title': '10반 - 26차시', 'date': '2026-06-09', 'class': '10반'},
    {'id': 327, 'title': '10반 - 27차시', 'date': '2026-06-11', 'class': '10반'},
    {'id': 328, 'title': '10반 - 28차시', 'date': '2026-06-15', 'class': '10반'},
    {'id': 329, 'title': '10반 - 29차시', 'date': '2026-06-16', 'class': '10반'},
    {'id': 330, 'title': '10반 - 30차시', 'date': '2026-06-18', 'class': '10반'},
    {'id': 331, 'title': '10반 - 31차시', 'date': '2026-06-22', 'class': '10반'},
    {'id': 332, 'title': '10반 - 32차시', 'date': '2026-06-23', 'class': '10반'},
    {'id': 333, 'title': '10반 - 33차시', 'date': '2026-06-25', 'class': '10반'},
    {'id': 334, 'title': '10반 - 34차시', 'date': '2026-06-29', 'class': '10반'},
    {'id': 335, 'title': '10반 - 35차시', 'date': '2026-06-30', 'class': '10반'},
    {'id': 336, 'title': '10반 - 36차시', 'date': '2026-07-02', 'class': '10반'},
    {'id': 337, 'title': '10반 - 37차시', 'date': '2026-07-06', 'class': '10반'},
    {'id': 338, 'title': '10반 - 38차시', 'date': '2026-07-07', 'class': '10반'},
    {'id': 339, 'title': '10반 - 39차시', 'date': '2026-07-09', 'class': '10반'},
    # 11반 (월수금)
    {'id': 401, 'title': '11반 - 1차시', 'date': '2026-04-13', 'class': '11반'},
    {'id': 402, 'title': '11반 - 2차시', 'date': '2026-04-15', 'class': '11반'},
    {'id': 403, 'title': '11반 - 3차시', 'date': '2026-04-17', 'class': '11반'},
    {'id': 404, 'title': '11반 - 4차시', 'date': '2026-04-20', 'class': '11반'},
    {'id': 405, 'title': '11반 - 5차시', 'date': '2026-04-22', 'class': '11반'},
    {'id': 406, 'title': '11반 - 6차시', 'date': '2026-04-24', 'class': '11반'},
    {'id': 407, 'title': '11반 - 7차시', 'date': '2026-04-27', 'class': '11반'},
    {'id': 408, 'title': '11반 - 8차시', 'date': '2026-04-29', 'class': '11반'},
    {'id': 409, 'title': '11반 - 9차시', 'date': '2026-05-01', 'class': '11반'},
    {'id': 410, 'title': '11반 - 10차시', 'date': '2026-05-04', 'class': '11반'},
    {'id': 411, 'title': '11반 - 11차시', 'date': '2026-05-06', 'class': '11반'},
    {'id': 412, 'title': '11반 - 12차시', 'date': '2026-05-08', 'class': '11반'},
    {'id': 413, 'title': '11반 - 13차시', 'date': '2026-05-11', 'class': '11반'},
    {'id': 414, 'title': '11반 - 14차시', 'date': '2026-05-13', 'class': '11반'},
    {'id': 415, 'title': '11반 - 15차시', 'date': '2026-05-15', 'class': '11반'},
    {'id': 416, 'title': '11반 - 16차시', 'date': '2026-05-18', 'class': '11반'},
    {'id': 417, 'title': '11반 - 17차시', 'date': '2026-05-20', 'class': '11반'},
    {'id': 418, 'title': '11반 - 18차시', 'date': '2026-05-22', 'class': '11반'},
    {'id': 419, 'title': '11반 - 19차시', 'date': '2026-05-25', 'class': '11반'},
    {'id': 420, 'title': '11반 - 20차시', 'date': '2026-05-27', 'class': '11반'},
    {'id': 421, 'title': '11반 - 21차시', 'date': '2026-05-29', 'class': '11반'},
    {'id': 422, 'title': '11반 - 22차시', 'date': '2026-06-01', 'class': '11반'},
    {'id': 423, 'title': '11반 - 23차시', 'date': '2026-06-03', 'class': '11반'},
    {'id': 424, 'title': '11반 - 24차시', 'date': '2026-06-05', 'class': '11반'},
    {'id': 425, 'title': '11반 - 25차시', 'date': '2026-06-08', 'class': '11반'},
    {'id': 426, 'title': '11반 - 26차시', 'date': '2026-06-10', 'class': '11반'},
    {'id': 427, 'title': '11반 - 27차시', 'date': '2026-06-12', 'class': '11반'},
    {'id': 428, 'title': '11반 - 28차시', 'date': '2026-06-15', 'class': '11반'},
    {'id': 429, 'title': '11반 - 29차시', 'date': '2026-06-17', 'class': '11반'},
    {'id': 430, 'title': '11반 - 30차시', 'date': '2026-06-19', 'class': '11반'},
    {'id': 431, 'title': '11반 - 31차시', 'date': '2026-06-22', 'class': '11반'},
    {'id': 432, 'title': '11반 - 32차시', 'date': '2026-06-24', 'class': '11반'},
    {'id': 433, 'title': '11반 - 33차시', 'date': '2026-06-26', 'class': '11반'},
    {'id': 434, 'title': '11반 - 34차시', 'date': '2026-06-29', 'class': '11반'},
    {'id': 435, 'title': '11반 - 35차시', 'date': '2026-07-01', 'class': '11반'},
    {'id': 436, 'title': '11반 - 36차시', 'date': '2026-07-03', 'class': '11반'},
    {'id': 437, 'title': '11반 - 37차시', 'date': '2026-07-06', 'class': '11반'},
    {'id': 438, 'title': '11반 - 38차시', 'date': '2026-07-08', 'class': '11반'},
    {'id': 439, 'title': '11반 - 39차시', 'date': '2026-07-10', 'class': '11반'},
    # 12반 (월화수)
    {'id': 501, 'title': '12반 - 1차시', 'date': '2026-04-13', 'class': '12반'},
    {'id': 502, 'title': '12반 - 2차시', 'date': '2026-04-14', 'class': '12반'},
    {'id': 503, 'title': '12반 - 3차시', 'date': '2026-04-15', 'class': '12반'},
    {'id': 504, 'title': '12반 - 4차시', 'date': '2026-04-20', 'class': '12반'},
    {'id': 505, 'title': '12반 - 5차시', 'date': '2026-04-21', 'class': '12반'},
    {'id': 506, 'title': '12반 - 6차시', 'date': '2026-04-22', 'class': '12반'},
    {'id': 507, 'title': '12반 - 7차시', 'date': '2026-04-27', 'class': '12반'},
    {'id': 508, 'title': '12반 - 8차시', 'date': '2026-04-28', 'class': '12반'},
    {'id': 509, 'title': '12반 - 9차시', 'date': '2026-04-29', 'class': '12반'},
    {'id': 510, 'title': '12반 - 10차시', 'date': '2026-05-04', 'class': '12반'},
    {'id': 511, 'title': '12반 - 11차시', 'date': '2026-05-06', 'class': '12반'},
    {'id': 512, 'title': '12반 - 12차시', 'date': '2026-05-07', 'class': '12반'},
    {'id': 513, 'title': '12반 - 13차시', 'date': '2026-05-11', 'class': '12반'},
    {'id': 514, 'title': '12반 - 14차시', 'date': '2026-05-12', 'class': '12반'},
    {'id': 515, 'title': '12반 - 15차시', 'date': '2026-05-13', 'class': '12반'},
    {'id': 516, 'title': '12반 - 16차시', 'date': '2026-05-18', 'class': '12반'},
    {'id': 517, 'title': '12반 - 17차시', 'date': '2026-05-19', 'class': '12반'},
    {'id': 518, 'title': '12반 - 18차시', 'date': '2026-05-20', 'class': '12반'},
    {'id': 519, 'title': '12반 - 19차시', 'date': '2026-05-25', 'class': '12반'},
    {'id': 520, 'title': '12반 - 20차시', 'date': '2026-05-26', 'class': '12반'},
    {'id': 521, 'title': '12반 - 21차시', 'date': '2026-05-27', 'class': '12반'},
    {'id': 522, 'title': '12반 - 22차시', 'date': '2026-06-01', 'class': '12반'},
    {'id': 523, 'title': '12반 - 23차시', 'date': '2026-06-02', 'class': '12반'},
    {'id': 524, 'title': '12반 - 24차시', 'date': '2026-06-03', 'class': '12반'},
    {'id': 525, 'title': '12반 - 25차시', 'date': '2026-06-08', 'class': '12반'},
    {'id': 526, 'title': '12반 - 26차시', 'date': '2026-06-09', 'class': '12반'},
    {'id': 527, 'title': '12반 - 27차시', 'date': '2026-06-10', 'class': '12반'},
    {'id': 528, 'title': '12반 - 28차시', 'date': '2026-06-15', 'class': '12반'},
    {'id': 529, 'title': '12반 - 29차시', 'date': '2026-06-16', 'class': '12반'},
    {'id': 530, 'title': '12반 - 30차시', 'date': '2026-06-17', 'class': '12반'},
    {'id': 531, 'title': '12반 - 31차시', 'date': '2026-06-22', 'class': '12반'},
    {'id': 532, 'title': '12반 - 32차시', 'date': '2026-06-23', 'class': '12반'},
    {'id': 533, 'title': '12반 - 33차시', 'date': '2026-06-24', 'class': '12반'},
    {'id': 534, 'title': '12반 - 34차시', 'date': '2026-06-29', 'class': '12반'},
    {'id': 535, 'title': '12반 - 35차시', 'date': '2026-06-30', 'class': '12반'},
    {'id': 536, 'title': '12반 - 36차시', 'date': '2026-07-01', 'class': '12반'},
    {'id': 537, 'title': '12반 - 37차시', 'date': '2026-07-06', 'class': '12반'},
    {'id': 538, 'title': '12반 - 38차시', 'date': '2026-07-07', 'class': '12반'},
    {'id': 539, 'title': '12반 - 39차시', 'date': '2026-07-08', 'class': '12반'},
    # 13반 (월화목)
    {'id': 601, 'title': '13반 - 1차시', 'date': '2026-04-13', 'class': '13반'},
    {'id': 602, 'title': '13반 - 2차시', 'date': '2026-04-14', 'class': '13반'},
    {'id': 603, 'title': '13반 - 3차시', 'date': '2026-04-16', 'class': '13반'},
    {'id': 604, 'title': '13반 - 4차시', 'date': '2026-04-20', 'class': '13반'},
    {'id': 605, 'title': '13반 - 5차시', 'date': '2026-04-21', 'class': '13반'},
    {'id': 606, 'title': '13반 - 6차시', 'date': '2026-04-23', 'class': '13반'},
    {'id': 607, 'title': '13반 - 7차시', 'date': '2026-04-27', 'class': '13반'},
    {'id': 608, 'title': '13반 - 8차시', 'date': '2026-04-28', 'class': '13반'},
    {'id': 609, 'title': '13반 - 9차시', 'date': '2026-04-30', 'class': '13반'},
    {'id': 610, 'title': '13반 - 10차시', 'date': '2026-05-04', 'class': '13반'},
    {'id': 611, 'title': '13반 - 11차시', 'date': '2026-05-06', 'class': '13반'},
    {'id': 612, 'title': '13반 - 12차시', 'date': '2026-05-07', 'class': '13반'},
    {'id': 613, 'title': '13반 - 13차시', 'date': '2026-05-11', 'class': '13반'},
    {'id': 614, 'title': '13반 - 14차시', 'date': '2026-05-12', 'class': '13반'},
    {'id': 615, 'title': '13반 - 15차시', 'date': '2026-05-14', 'class': '13반'},
    {'id': 616, 'title': '13반 - 16차시', 'date': '2026-05-18', 'class': '13반'},
    {'id': 617, 'title': '13반 - 17차시', 'date': '2026-05-19', 'class': '13반'},
    {'id': 618, 'title': '13반 - 18차시', 'date': '2026-05-21', 'class': '13반'},
    {'id': 619, 'title': '13반 - 19차시', 'date': '2026-05-25', 'class': '13반'},
    {'id': 620, 'title': '13반 - 20차시', 'date': '2026-05-26', 'class': '13반'},
    {'id': 621, 'title': '13반 - 21차시', 'date': '2026-05-28', 'class': '13반'},
    {'id': 622, 'title': '13반 - 22차시', 'date': '2026-06-01', 'class': '13반'},
    {'id': 623, 'title': '13반 - 23차시', 'date': '2026-06-02', 'class': '13반'},
    {'id': 624, 'title': '13반 - 24차시', 'date': '2026-06-04', 'class': '13반'},
    {'id': 625, 'title': '13반 - 25차시', 'date': '2026-06-08', 'class': '13반'},
    {'id': 626, 'title': '13반 - 26차시', 'date': '2026-06-09', 'class': '13반'},
    {'id': 627, 'title': '13반 - 27차시', 'date': '2026-06-11', 'class': '13반'},
    {'id': 628, 'title': '13반 - 28차시', 'date': '2026-06-15', 'class': '13반'},
    {'id': 629, 'title': '13반 - 29차시', 'date': '2026-06-16', 'class': '13반'},
    {'id': 630, 'title': '13반 - 30차시', 'date': '2026-06-18', 'class': '13반'},
    {'id': 631, 'title': '13반 - 31차시', 'date': '2026-06-22', 'class': '13반'},
    {'id': 632, 'title': '13반 - 32차시', 'date': '2026-06-23', 'class': '13반'},
    {'id': 633, 'title': '13반 - 33차시', 'date': '2026-06-25', 'class': '13반'},
    {'id': 634, 'title': '13반 - 34차시', 'date': '2026-06-29', 'class': '13반'},
    {'id': 635, 'title': '13반 - 35차시', 'date': '2026-06-30', 'class': '13반'},
    {'id': 636, 'title': '13반 - 36차시', 'date': '2026-07-02', 'class': '13반'},
    {'id': 637, 'title': '13반 - 37차시', 'date': '2026-07-06', 'class': '13반'},
    {'id': 638, 'title': '13반 - 38차시', 'date': '2026-07-07', 'class': '13반'},
    {'id': 639, 'title': '13반 - 39차시', 'date': '2026-07-09', 'class': '13반'},
]

def index(request):
    # 활성 퀴즈가 있는 차시 번호 집합
    active_chapters = set(
        QuizSet.objects.filter(is_active=True).values_list('chapter_num', flat=True)
    )
    lessons_by_class = OrderedDict()
    for lesson in LESSONS:
        c = lesson['class']
        if c not in lessons_by_class:
            lessons_by_class[c] = []
        entry = dict(lesson)
        entry['has_quiz'] = (lesson['id'] % 100) in active_chapters
        lessons_by_class[c].append(entry)
    return render(request, 'feedback/index.html', {'lessons_by_class': lessons_by_class})

def feedback_create(request, lesson_id):
    lesson = next((l for l in LESSONS if l['id'] == lesson_id), None)
    if not lesson:
        return redirect('index')

    if request.method == 'POST':
        student_id   = request.POST['student_id'].strip()
        student_num  = request.POST['student_num'].strip()
        student_name = request.POST['student_name'].strip()
        summary      = request.POST['summary'].strip()
        problem      = request.POST['problem'].strip()
        career       = request.POST['career'].strip()
        deeplearn    = request.POST['deeplearn'].strip()
        peer         = request.POST['peer'].strip()

        # 빈 칸 서버측 검증
        if not all([student_id, student_num, student_name,
                    summary, problem, career, deeplearn, peer]):
            return render(request, 'feedback/create.html', {
                'lesson': lesson,
                'error': '모든 항목을 입력해주세요!',
                'prev': request.POST,
                'is_edit': False,
            })

        # ── 반 검증: 학번 앞 3자리와 수업 반이 일치하는지 확인 ──
        if student_id in _VALID_IDS:
            student_class_code = student_id[:3]
            lesson_class_code  = _LESSON_CLASS_CODE.get(lesson_id // 100)
            if lesson_class_code and student_class_code != lesson_class_code:
                # 올바른 차시 계산 (같은 차시, 본인 반)
                chapter_num    = lesson_id % 100
                correct_prefix = next(
                    (k for k, v in _LESSON_CLASS_CODE.items() if v == student_class_code), None
                )
                correct_lesson_id = (correct_prefix * 100 + chapter_num) if correct_prefix else None
                return render(request, 'feedback/create.html', {
                    'lesson': lesson,
                    'class_error': True,
                    'student_class_name': _CLASS_CODE_NAME.get(student_class_code, ''),
                    'lesson_class_name':  _CLASS_CODE_NAME.get(lesson_class_code, ''),
                    'correct_lesson_id':  correct_lesson_id,
                    'chapter_num': chapter_num,
                    'prev': request.POST,
                    'is_edit': False,
                })

        # 최소 글자수 검증
        if any(len(x) < 2 for x in [summary, problem, career, deeplearn, peer]):
            return render(request, 'feedback/create.html', {
                'lesson': lesson,
                'error': '각 항목을 좀 더 자세히 입력해주세요! (최소 2자 이상)',
                'prev': request.POST,
                'is_edit': False,
            })

        already = False
        try:
            sheet = get_sheet()
            records = sheet.get_all_values()
            for row in records[1:]:
                # 실제 시트 레이아웃: [1]=수업명, [2]=학번
                if len(row) >= 3 and row[1] == lesson['title'] and row[2] == student_id:
                    already = True
                    break
        except Exception as e:
            print(f"중복 확인 오류: {e}")

        if not already:
            def save_to_sheet():
                try:
                    sheet = get_sheet()
                    sheet.append_row([
                        datetime.now().strftime('%Y-%m-%d %H:%M'),
                        lesson['title'],
                        student_id,    # C[2]: 학번
                        student_num,   # D[3]: 번호
                        student_name,  # E[4]: 이름
                        summary,       # F[5]: 핵심개념
                        problem,       # G[6]: 문제해결
                        career,        # H[7]: 진로연결
                        deeplearn,     # I[8]: 심화학습
                        peer,          # J[9]: 칭찬
                    ])
                except Exception as e:
                    print(f"구글 시트 저장 오류: {e}")
            threading.Thread(target=save_to_sheet).start()

        return render(request, 'feedback/done.html', {
            'lesson': lesson,
            'already': already,
        })

    return render(request, 'feedback/create.html', {
        'lesson': lesson,
        'prev': {},
        'is_edit': False,
        'error': None,
    })


def feedback_edit(request, lesson_id, student_id):
    lesson = next((l for l in LESSONS if l['id'] == lesson_id), None)
    if not lesson:
        return redirect('index')

    # 기존 데이터 불러오기
    existing = {}
    row_index = None
    try:
        sheet = get_sheet()
        records = sheet.get_all_values()
        for i, row in enumerate(records[1:], start=2):
            # 실제 시트: [1]=수업명, [2]=학번
            if len(row) >= 3 and row[1] == lesson['title'] and row[2] == student_id:
                existing = {
                    'student_id':   row[2],
                    'student_num':  row[3],
                    'student_name': row[4] if len(row) > 4 else '',
                    'summary':      row[5] if len(row) > 5 else '',
                    'problem':      row[6] if len(row) > 6 else '',
                    'career':       row[7] if len(row) > 7 else '',
                    'deeplearn':    row[8] if len(row) > 8 else '',
                    'peer':         row[9] if len(row) > 9 else '',
                }
                row_index = i
                break
    except Exception as e:
        print(f"데이터 불러오기 오류: {e}")

    if not existing:
        return redirect('index')

    if request.method == 'POST':
        summary   = request.POST['summary'].strip()
        problem   = request.POST['problem'].strip()
        career    = request.POST['career'].strip()
        deeplearn = request.POST['deeplearn'].strip()
        peer      = request.POST['peer'].strip()

        if any(len(x) < 2 for x in [summary, problem, career, deeplearn, peer]):
            return render(request, 'feedback/create.html', {
                'lesson': lesson,
                'error': '각 항목을 좀 더 자세히 입력해주세요!',
                'prev': request.POST,
                'is_edit': True,
            })

        # 구글 시트 해당 행 수정
        def update_sheet():
            try:
                sheet = get_sheet()
                # 실제 시트: 핵심개념=F[5], 문제해결=G[6], ..., 칭찬=J[9]
                sheet.update(f'F{row_index}:J{row_index}', [[
                    summary, problem, career, deeplearn, peer
                ]])
                # A열 수정시간 업데이트
                sheet.update(f'A{row_index}', [[
                    datetime.now().strftime('%Y-%m-%d %H:%M') + ' (수정)'
                ]])
            except Exception as e:
                print(f"수정 오류: {e}")
        threading.Thread(target=update_sheet).start()

        return render(request, 'feedback/done.html', {
            'lesson': lesson,
            'already': False,
            'is_edit': True,
        })

    return render(request, 'feedback/create.html', {
        'lesson': lesson,
        'prev': existing,
        'is_edit': True,
        'error': None,
    })

def lesson_result(request, lesson_id):
    lesson = next((l for l in LESSONS if l['id'] == lesson_id), None)
    if not lesson:
        return redirect('index')

    student_id = request.GET.get('student_id', '').strip()
    record = None
    error  = None

    if student_id:
        try:
            sheet = get_sheet()
            records = sheet.get_all_values()
            for row in records[1:]:
                # 실제 시트: [1]=수업명, [2]=학번
                if len(row) >= 3 and row[1] == lesson['title'] and row[2] == student_id:
                    record = {
                        'student_id':   row[2],
                        'student_num':  row[3],
                        'student_name': row[4] if len(row) > 4 else '',
                        'summary':      row[5] if len(row) > 5 else '',
                        'problem':      row[6] if len(row) > 6 else '',
                        'career':       row[7] if len(row) > 7 else '',
                        'deeplearn':    row[8] if len(row) > 8 else '',
                        'peer':         row[9] if len(row) > 9 else '',
                    }
                    break
            if not record:
                error = "해당 학번의 제출 기록이 없습니다."
        except Exception as e:
            error = "조회 중 오류가 발생했습니다."

    return render(request, 'feedback/result.html', {
        'lesson': lesson,
        'record': record,
        'error': error,
        'student_id': student_id,
    })

def student_summary(request, student_id):
    try:
        sheet = get_sheet()
        records = sheet.get_all_values()
        student_records = []
        student_name = ''
        for row in records[1:]:
            # 실제 시트: [2]=학번, [4]=이름
            if len(row) >= 3 and row[2] == student_id:
                student_name = row[4] if len(row) > 4 else ''
                student_records.append({
                    'date':         row[0],
                    'lesson_title': row[1],
                    'summary':      row[5] if len(row) > 5 else '',
                    'problem':      row[6] if len(row) > 6 else '',
                    'career':       row[7] if len(row) > 7 else '',
                    'deeplearn':    row[8] if len(row) > 8 else '',
                    'peer':         row[9] if len(row) > 9 else '',
                })
    except Exception as e:
        student_records = []
        student_name = ''

    if not student_records:
        return render(request, 'feedback/not_found.html')

    return render(request, 'feedback/student_summary.html', {
        'records':      student_records,
        'student_name': student_name,
        'student_id':   student_id,
        'count':        len(student_records),
    })


# ═══════════════════════════════════════════════════════════════
#  활동문제 채점 시스템  (Quiz System)
# ═══════════════════════════════════════════════════════════════

TEACHER_PASSWORD = os.environ.get('TEACHER_PASSWORD', 'info1234')


# ── 헬퍼: 이미지 리사이즈 → base64 ──────────────────────────────

def _resize_image_to_b64(image_bytes):
    """업로드 이미지를 최대 1200px JPEG(quality 70%)로 변환 후 base64 반환"""
    try:
        from PIL import Image
        img = Image.open(io.BytesIO(image_bytes))
        if img.mode in ('RGBA', 'P', 'LA'):
            img = img.convert('RGB')
        max_px = 1200
        if max(img.size) > max_px:
            ratio = max_px / max(img.size)
            img = img.resize((int(img.width * ratio), int(img.height * ratio)), Image.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format='JPEG', quality=70)
        return base64.b64encode(buf.getvalue()).decode('utf-8')
    except Exception as e:
        print(f"이미지 처리 오류: {e}")
        return base64.b64encode(image_bytes).decode('utf-8')


# ── 헬퍼: PDF → 텍스트 ───────────────────────────────────────────

def _extract_text_from_pdf(pdf_bytes):
    """pdfplumber로 PDF 텍스트 추출"""
    try:
        import pdfplumber
        parts = []
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    parts.append(t)
        return '\n\n'.join(parts)
    except Exception as e:
        print(f"PDF 추출 오류: {e}")
        return ''


# ── 헬퍼: Claude로 지도서 → 문제/정답 JSON 추출 ──────────────────

def _extract_questions_with_claude(guide_text, chapter_num):
    """지도서 텍스트를 Claude에게 보내 문제·정답 구조화"""
    api_key = os.environ.get('ANTHROPIC_API_KEY', '')
    if not api_key:
        return {'title': f'{chapter_num}차시', 'questions': []}
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        prompt = f"""다음은 정보 교과 {chapter_num}차시 교사용 지도서 내용입니다.
학생 활동문제(연습문제·탐구·확인문제 등)와 그 모범답안을 모두 추출해 주세요.

지도서 내용:
{guide_text[:8000]}

아래 JSON 형식으로만 응답하세요 (설명 없이):
{{
  "title": "이 차시 학습 주제 제목",
  "questions": [
    {{
      "number": 1,
      "content": "학생 교과서에 나오는 문제 전문",
      "model_answer": "교사 지도서 모범답안(상세히)",
      "score": 10
    }}
  ]
}}
문제가 없으면 questions 를 빈 배열로 반환하세요."""
        msg = client.messages.create(
            model='claude-opus-4-6',
            max_tokens=4096,
            messages=[{'role': 'user', 'content': prompt}],
        )
        raw = msg.content[0].text
        m = re.search(r'\{[\s\S]*\}', raw)
        if m:
            return json.loads(m.group())
    except Exception as e:
        print(f"Claude 문제 추출 오류: {e}")
    return {'title': f'{chapter_num}차시', 'questions': []}


# ── 헬퍼: Claude로 학생 답안 채점 ────────────────────────────────

def _grade_with_claude(q_content, model_answer, ans_text, ans_code, ans_img_b64, max_score):
    """학생 답안을 Claude Haiku로 채점 → (score, feedback) 반환"""
    api_key = os.environ.get('ANTHROPIC_API_KEY', '')
    if not api_key:
        return None, '자동 채점 미설정 – 교사가 직접 채점합니다.'
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)

        prompt = f"""당신은 친절하고 꼼꼼한 정보 교과 교사입니다.

[문제]
{q_content}

[모범 답안]
{model_answer}

[배점] {max_score}점
"""
        if ans_text:
            prompt += f"\n[학생 텍스트 답안]\n{ans_text}\n"
        if ans_code:
            prompt += f"\n[학생 코드 답안]\n```\n{ans_code}\n```\n"
        if ans_img_b64:
            prompt += "\n[학생 사진 답안: 아래 첨부 이미지 참고]\n"

        prompt += f"""
위 답안을 채점하고 **JSON 형식으로만** 응답하세요:
{{"score": (0~{max_score} 정수), "feedback": "잘한 점과 개선할 점을 포함한 2~3문장 한국어 피드백"}}"""

        content = [{'type': 'text', 'text': prompt}]
        if ans_img_b64:
            content.append({
                'type': 'image',
                'source': {'type': 'base64', 'media_type': 'image/jpeg', 'data': ans_img_b64},
            })

        msg = client.messages.create(
            model='claude-haiku-4-5-20251001',
            max_tokens=400,
            messages=[{'role': 'user', 'content': content}],
        )
        raw = msg.content[0].text
        m = re.search(r'\{[^{}]*\}', raw)
        if m:
            r = json.loads(m.group())
            score = max(0, min(max_score, int(r.get('score', 0))))
            return score, r.get('feedback', '')
    except Exception as e:
        print(f"Claude 채점 오류: {e}")
    return None, '자동 채점 중 오류 발생 – 교사가 직접 채점합니다.'


# ── View: 교사 – 지도서 1회 업로드 ──────────────────────────────────

def guide_upload(request):
    """교사용: 지도서 텍스트를 입력받아 저장 (텍스트 전용 – PDF 업로드 제거)"""
    error = success = None
    guides = GuideBook.objects.all().order_by('-uploaded_at')

    if request.method == 'POST':
        pw = request.POST.get('password', '')
        if pw != TEACHER_PASSWORD:
            error = '비밀번호가 올바르지 않습니다.'
        else:
            name        = request.POST.get('guide_name', '').strip() or '정보 교과 지도서'
            manual_text = request.POST.get('manual_text', '').strip()

            if not manual_text:
                error = '지도서 내용을 입력해주세요.'
            elif len(manual_text) < 50:
                error = '내용이 너무 짧습니다. 지도서 내용을 더 많이 입력해주세요.'
            else:
                try:
                    gb = GuideBook.objects.create(
                        name=name,
                        full_text=manual_text,
                        page_count=0,
                    )
                    success = f'✓ 지도서 [{gb.name}] 저장 완료! ({len(manual_text):,}자)'
                    guides = GuideBook.objects.all().order_by('-uploaded_at')
                except Exception as e:
                    error = f'저장 중 오류: {e}'

    return render(request, 'feedback/guide_upload.html', {
        'error': error, 'success': success, 'guides': guides,
    })


# ── View: 교사 – 범위 설정 → 문제 생성 ──────────────────────────────

def quiz_generate(request):
    """교사용: 저장된 지도서에서 교과서 범위를 지정해 문제 생성"""
    error = success = None
    guides   = GuideBook.objects.all().order_by('-uploaded_at')
    existing = QuizSet.objects.all().order_by('chapter_num')

    if request.method == 'POST':
        pw = request.POST.get('password', '')
        if pw != TEACHER_PASSWORD:
            error = '비밀번호가 올바르지 않습니다.'
        else:
            guide_id    = request.POST.get('guide_id', '').strip()
            chapter_raw = request.POST.get('chapter_num', '').strip()
            topic       = request.POST.get('range_topic', '').strip()
            pages       = request.POST.get('range_pages', '').strip()
            extra_text  = request.POST.get('extra_text', '').strip()
            q_count     = request.POST.get('q_count', '3').strip()
            q_type      = request.POST.get('q_type', '혼합').strip()

            if not chapter_raw:
                error = '차시 번호를 입력해주세요.'
            elif not topic:
                error = '교과서 범위(주제/단원명)를 입력해주세요.'
            else:
                try:
                    chapter_num = int(chapter_raw)
                    n_questions = max(1, min(10, int(q_count)))

                    # 범위 텍스트 추출
                    range_text = ''
                    guide = None
                    if guide_id:
                        try:
                            guide = GuideBook.objects.get(id=int(guide_id))
                            full = guide.full_text
                            # 페이지 범위 파싱 (예: "45-62" 또는 "45")
                            if pages:
                                p_parts = pages.replace('~', '-').split('-')
                                try:
                                    p_start = int(p_parts[0]) - 1
                                    p_end   = int(p_parts[-1]) if len(p_parts) > 1 else p_start + 30
                                except Exception:
                                    p_start, p_end = 0, 999
                                # 텍스트를 페이지 단위로 분할 (\f 또는 단순 균등 분할)
                                pages_split = full.split('\f')
                                if len(pages_split) > 3:
                                    range_text = '\n'.join(pages_split[p_start:p_end])
                                else:
                                    # 균등 분할 추정
                                    total_chars = len(full)
                                    if guide.page_count > 0:
                                        cpp = total_chars // guide.page_count
                                        range_text = full[p_start * cpp : p_end * cpp]
                                    else:
                                        range_text = full
                            else:
                                # 주제 키워드로 관련 구간 추출 (전후 5000자)
                                idx = full.find(topic[:10])
                                if idx >= 0:
                                    range_text = full[max(0, idx-500) : idx+8000]
                                else:
                                    range_text = full[:12000]
                        except GuideBook.DoesNotExist:
                            range_text = ''

                    # extra_text 보완
                    if extra_text:
                        range_text = (range_text + '\n\n' + extra_text).strip()

                    if len(range_text) < 30 and not extra_text:
                        error = '지도서를 선택하거나 교과서 내용을 직접 입력해주세요.'
                    else:
                        result = _generate_questions_with_claude(
                            range_text, chapter_num, topic, n_questions, q_type
                        )
                        qs, created = QuizSet.objects.update_or_create(
                            chapter_num=chapter_num,
                            defaults={
                                'guidebook':   guide,
                                'title':       result.get('title', f'{chapter_num}차시 – {topic}'),
                                'range_topic': topic,
                                'range_pages': pages,
                                'range_text':  range_text[:5000],
                                'is_active':   True,
                            }
                        )
                        qs.questions.all().delete()
                        for q in result.get('questions', []):
                            Question.objects.create(
                                quizset=qs,
                                number=q.get('number', 1),
                                content=q.get('content', ''),
                                model_answer=q.get('model_answer', ''),
                                score=q.get('score', 10),
                            )
                        cnt = qs.questions.count()
                        action = '생성' if created else '재생성'
                        success = f'✓ [{qs.title}] 문제 {cnt}개 {action} 완료!'
                        existing = QuizSet.objects.all().order_by('chapter_num')
                except Exception as e:
                    error = f'문제 생성 중 오류: {e}'

    return render(request, 'feedback/quiz_generate.html', {
        'error': error, 'success': success,
        'guides': guides, 'existing': existing,
    })


def _generate_questions_with_claude(range_text, chapter_num, topic, n_questions, q_type):
    """Claude로 교과서 범위 텍스트 → 문제·정답 생성"""
    api_key = os.environ.get('ANTHROPIC_API_KEY', '')
    if not api_key:
        return {'title': f'{chapter_num}차시 – {topic}', 'questions': []}
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)

        type_guide = {
            '단답형':  '모두 단답형(짧은 답변) 문제로 만드세요.',
            '서술형':  '모두 서술형(2~4문장 설명) 문제로 만드세요.',
            '코딩':    '모두 코드 작성 또는 코드 분석 문제로 만드세요.',
            '혼합':    '단답형, 서술형, 코딩 문제를 고루 섞어 만드세요.',
        }.get(q_type, '단답형, 서술형을 고루 섞어 만드세요.')

        prompt = f"""당신은 고등학교 정보 교과 교사입니다.
아래는 교과서 [{topic}] 단원(또는 {chapter_num}차시)의 내용입니다.

--- 교과서 내용 ---
{range_text[:10000]}
---

위 내용을 바탕으로 학생 활동 문제 {n_questions}개를 만들어 주세요.
{type_guide}

요구사항:
- 교과서 내용에서 직접 출제하세요 (내용 범위를 벗어나지 마세요)
- 학생이 이해했는지 확인할 수 있는 핵심 문제로 만드세요
- 모범답안은 교사가 채점할 수 있을 만큼 상세하게 작성하세요
- 배점은 문제 난이도에 따라 5~20점 사이로 설정하세요

아래 JSON 형식으로만 응답하세요 (설명 없이):
{{
  "title": "{chapter_num}차시 – {{단원/주제 제목}}",
  "questions": [
    {{
      "number": 1,
      "type": "단답형|서술형|코딩 중 하나",
      "content": "학생에게 보여줄 문제 전문",
      "model_answer": "교사용 모범답안 (상세히)",
      "score": 10
    }}
  ]
}}"""

        msg = client.messages.create(
            model='claude-sonnet-4-20250514',
            max_tokens=4096,
            messages=[{'role': 'user', 'content': prompt}],
        )
        raw = msg.content[0].text
        m = re.search(r'\{[\s\S]*\}', raw)
        if m:
            return json.loads(m.group())
    except Exception as e:
        print(f'Claude 문제 생성 오류: {e}')
    return {'title': f'{chapter_num}차시 – {topic}', 'questions': []}


# ── View: 학생 – 문제 풀기 ────────────────────────────────────────

def quiz_solve(request, lesson_id):
    """학생용: 학번 입력 → 문제 제시 → 답안 작성"""
    lesson = next((l for l in LESSONS if l['id'] == lesson_id), None)
    if not lesson:
        return redirect('index')

    chapter_num = lesson_id % 100
    quizset = QuizSet.objects.filter(chapter_num=chapter_num, is_active=True).first()
    if not quizset or not quizset.questions.exists():
        return render(request, 'feedback/quiz_no_quiz.html', {'lesson': lesson})

    # 이미 제출했으면 결과 페이지로
    student_id = request.GET.get('student_id', '').strip()
    if student_id:
        if StudentAnswer.objects.filter(
            question__quizset=quizset, lesson_id=lesson_id, student_id=student_id
        ).exists():
            return redirect(f'/quiz/{lesson_id}/result/?student_id={student_id}')

    return render(request, 'feedback/quiz_solve.html', {
        'lesson': lesson,
        'quizset': quizset,
        'questions': quizset.questions.all(),
        'student_id': student_id,
    })


# ── View: 학생 – 답안 제출 ────────────────────────────────────────

def quiz_submit(request, lesson_id):
    """학생 답안을 저장하고 Claude 비동기 채점 시작"""
    lesson = next((l for l in LESSONS if l['id'] == lesson_id), None)
    if not lesson or request.method != 'POST':
        return redirect('index')

    chapter_num = lesson_id % 100
    quizset = QuizSet.objects.filter(chapter_num=chapter_num, is_active=True).first()
    if not quizset:
        return redirect('index')

    student_id   = request.POST.get('student_id', '').strip()
    student_num  = request.POST.get('student_num', '').strip()
    student_name = request.POST.get('student_name', '').strip()

    if not all([student_id, student_num, student_name]):
        return redirect('quiz_solve', lesson_id=lesson_id)

    # 중복 제출 방지
    if StudentAnswer.objects.filter(
        question__quizset=quizset, lesson_id=lesson_id, student_id=student_id
    ).exists():
        return redirect(f'/quiz/{lesson_id}/result/?student_id={student_id}')

    saved = []
    for q in quizset.questions.all():
        ans_text = request.POST.get(f'ans_text_{q.id}', '').strip()
        ans_code = request.POST.get(f'ans_code_{q.id}', '').strip()
        ans_img_b64 = ''
        img_file = request.FILES.get(f'ans_img_{q.id}')
        if img_file:
            try:
                ans_img_b64 = _resize_image_to_b64(img_file.read())
            except Exception as e:
                print(f"이미지 처리 오류: {e}")

        if not any([ans_text, ans_code, ans_img_b64]):
            continue  # 빈 답안 건너뜀

        sa = StudentAnswer.objects.create(
            question=q,
            lesson_id=lesson_id,
            student_id=student_id,
            student_num=student_num,
            student_name=student_name,
            answer_text=ans_text,
            answer_code=ans_code,
            answer_image=ans_img_b64,
            max_score=q.score,
        )
        saved.append(sa)

        # ── 비동기 Claude 채점 ──
        def do_grade(sa=sa, q=q):
            score, fb = _grade_with_claude(
                q.content, q.model_answer,
                sa.answer_text, sa.answer_code, sa.answer_image,
                q.score,
            )
            sa.score = score
            sa.ai_feedback = fb
            sa.save()
        threading.Thread(target=do_grade, daemon=True).start()

    return render(request, 'feedback/quiz_submitted.html', {
        'lesson': lesson,
        'quizset': quizset,
        'student_name': student_name,
        'student_id': student_id,
        'answer_count': len(saved),
    })


# ── View: 학생 – 채점 결과 조회 ───────────────────────────────────

def quiz_result(request, lesson_id):
    """학번으로 채점 결과 조회"""
    lesson = next((l for l in LESSONS if l['id'] == lesson_id), None)
    if not lesson:
        return redirect('index')

    student_id = request.GET.get('student_id', '').strip()
    if not student_id:
        # 학번 입력 폼 표시
        return render(request, 'feedback/quiz_result.html', {
            'lesson': lesson, 'need_id': True,
        })

    chapter_num = lesson_id % 100
    quizset = QuizSet.objects.filter(chapter_num=chapter_num).first()
    answers = []
    if quizset:
        answers = list(
            StudentAnswer.objects.filter(
                question__quizset=quizset,
                lesson_id=lesson_id,
                student_id=student_id,
            ).select_related('question').order_by('question__number')
        )

    if not answers:
        return render(request, 'feedback/quiz_result.html', {
            'lesson': lesson, 'not_found': True, 'student_id': student_id,
        })

    graded     = [a for a in answers if a.score is not None]
    total      = sum(a.score for a in graded)
    max_total  = sum(a.question.score for a in answers)
    all_graded = len(graded) == len(answers)

    return render(request, 'feedback/quiz_result.html', {
        'lesson': lesson,
        'quizset': quizset,
        'answers': answers,
        'student_name': answers[0].student_name,
        'student_id': student_id,
        'total': total,
        'max_total': max_total,
        'all_graded': all_graded,
        'need_id': False,
        'not_found': False,
    })


# ── View: 교사 대시보드 ───────────────────────────────────────────

def teacher_dashboard(request):
    """교사 로그인 + 퀴즈 관리 목록 + 피드백 제출 현황"""
    if request.method == 'POST':
        pw = request.POST.get('password', '')
        if pw == TEACHER_PASSWORD:
            request.session['teacher_auth'] = True
        else:
            return render(request, 'feedback/teacher_dashboard.html',
                          {'error': '비밀번호가 올바르지 않습니다.'})

    if not request.session.get('teacher_auth'):
        return render(request, 'feedback/teacher_dashboard.html', {})

    quizsets = QuizSet.objects.prefetch_related('questions').order_by('chapter_num')
    stats = []
    for qs in quizsets:
        sub = StudentAnswer.objects.filter(question__quizset=qs)
        unconfirmed = sub.filter(is_confirmed=False, score__isnull=False).count()
        pending_grade = sub.filter(score__isnull=True).count()
        total_students = sub.values('student_id', 'lesson_id').distinct().count()
        stats.append({
            'qs': qs,
            'total_students': total_students,
            'unconfirmed': unconfirmed,
            'pending_grade': pending_grade,
        })

    # ── 피드백 제출 현황 (Google Sheets 기반) ──
    feedback_stats = []
    feedback_error = None
    try:
        sheet = get_sheet()
        rows = sheet.get_all_values()
        data_rows = rows[1:] if len(rows) > 1 else []

        # 반별 × 차시별 제출자 집계
        from collections import defaultdict
        import re as _re
        # class_prefix: 1→8반, 2→9반, ...
        prefix_map = {1:'8반', 2:'9반', 3:'10반', 4:'11반', 5:'12반', 6:'13반'}

        # 오늘 날짜 기준 지난 차시 중 반별 최근 5개 제출 현황
        today_str = datetime.now().strftime('%Y-%m-%d')

        # 반별 차시별 제출자 수
        # 실제 시트: [1]=수업명, [2]=학번
        class_lesson_students = defaultdict(lambda: defaultdict(set))
        for row in data_rows:
            if len(row) < 3:
                continue
            title          = row[1]
            student_id_val = row[2]
            if not title or not student_id_val:
                continue
            lesson_obj = _TITLE_TO_LESSON.get(title)
            if not lesson_obj:
                continue
            lid        = lesson_obj['id']
            class_name = lesson_obj['class']
            class_lesson_students[class_name][lid].add(student_id_val)

        # 반별로 최근 차시 현황 정리
        today_dt = datetime.now().date()
        for class_name in ['8반','9반','10반','11반','12반','13반']:
            lessons_for_class = [l for l in LESSONS if l['class'] == class_name]
            # 오늘 이전 차시만 (이미 진행된 수업)
            past_lessons = [l for l in lessons_for_class
                            if l['date'] <= today_str]
            if not past_lessons:
                continue
            # 최근 5차시
            recent = past_lessons[-5:]
            lesson_data = []
            for l in recent:
                submitted = len(class_lesson_students[class_name].get(l['id'], set()))
                lesson_data.append({
                    'title': l['title'],
                    'date': l['date'],
                    'submitted': submitted,
                })
            feedback_stats.append({
                'class_name': class_name,
                'lessons': lesson_data,
                'total_submitted': sum(len(v) for v in class_lesson_students[class_name].values()),
            })
    except Exception as e:
        feedback_error = f'피드백 현황 로드 오류: {e}'

    return render(request, 'feedback/teacher_dashboard.html', {
        'authenticated': True,
        'stats': stats,
        'feedback_stats': feedback_stats,
        'feedback_error': feedback_error,
    })


# ── View: 교사 – QuizSet 제출 현황 ───────────────────────────────

def teacher_quiz_detail(request, quizset_id):
    if not request.session.get('teacher_auth'):
        return redirect('teacher_dashboard')

    quizset   = get_object_or_404(QuizSet, id=quizset_id)
    questions = list(quizset.questions.all())
    answers   = (StudentAnswer.objects
                 .filter(question__quizset=quizset)
                 .select_related('question')
                 .order_by('lesson_id', 'student_num'))

    # 학생별 집계
    lesson_map = {l['id']: l['title'] for l in LESSONS}
    students_by_lesson = {}
    for ans in answers:
        key = (ans.lesson_id, ans.student_id)
        lid = ans.lesson_id
        if lid not in students_by_lesson:
            students_by_lesson[lid] = {}
        if ans.student_id not in students_by_lesson[lid]:
            students_by_lesson[lid][ans.student_id] = {
                'student_num':  ans.student_num,
                'student_name': ans.student_name,
                'student_id':   ans.student_id,
                'answers':      {},
                'total': 0, 'max': 0,
                'all_confirmed': True,
            }
        d = students_by_lesson[lid][ans.student_id]
        d['answers'][ans.question.number] = ans
        if ans.score is not None:
            d['total'] += ans.score
        d['max'] += ans.question.score
        if not ans.is_confirmed:
            d['all_confirmed'] = False

    lessons_data = {
        lid: {
            'title': lesson_map.get(lid, str(lid)),
            'students': sorted(smap.values(), key=lambda x: x['student_num']),
        }
        for lid, smap in sorted(students_by_lesson.items())
    }

    return render(request, 'feedback/teacher_quiz_detail.html', {
        'quizset':      quizset,
        'questions':    questions,
        'lessons_data': lessons_data,
    })


# ── View: 교사 – 개별 학생 채점 ──────────────────────────────────

def teacher_grade_student(request, quizset_id, lesson_id, student_id):
    if not request.session.get('teacher_auth'):
        return redirect('teacher_dashboard')

    quizset = get_object_or_404(QuizSet, id=quizset_id)
    answers = list(
        StudentAnswer.objects.filter(
            question__quizset=quizset,
            lesson_id=lesson_id,
            student_id=student_id,
        ).select_related('question').order_by('question__number')
    )
    if not answers:
        return redirect('teacher_quiz_detail', quizset_id=quizset_id)

    lesson = next((l for l in LESSONS if l['id'] == int(lesson_id)), None)

    if request.method == 'POST':
        for ans in answers:
            raw_score = request.POST.get(f'score_{ans.id}', '')
            raw_fb    = request.POST.get(f'feedback_{ans.id}', '').strip()
            if raw_score != '':
                try:
                    ans.score = max(0, min(ans.question.score, int(raw_score)))
                except ValueError:
                    pass
            ans.teacher_feedback = raw_fb
            ans.is_confirmed = True
            ans.save()
        return redirect('teacher_quiz_detail', quizset_id=quizset_id)

    return render(request, 'feedback/teacher_grade_student.html', {
        'quizset':      quizset,
        'answers':      answers,
        'lesson':       lesson,
        'student_id':   student_id,
        'student_name': answers[0].student_name,
    })


# ═══════════════════════════════════════════════════════════════
#  학생 피드백 데이터 리포트  (Report System)
# ═══════════════════════════════════════════════════════════════

# ── 구글 시트 실제 컬럼 레이아웃 ──────────────────────────────────
# [0] 날짜  [1] 수업명  [2] 학번(student_id)  [3] 번호(student_num)
# [4] 이름  [5] 핵심개념  [6] 문제해결  [7] 진로연결  [8] 심화학습  [9] 칭찬

# 수업명에서 반 이름 파싱: "8반 - 3차시" → "8반"
_CLASS_NAME_TO_CODE = {
    '8반': '108', '9반': '109', '10반': '110',
    '11반': '111', '12반': '112', '13반': '113',
}

def _class_from_title(title):
    """수업명 문자열에서 반 이름 반환. 예: '8반 - 3차시' → '8반'"""
    for name in _CLASS_NAME_TO_CODE:
        if name in title:
            return name
    return '알 수 없음'

def _class_code_from_title(title):
    """수업명에서 반 코드 반환. 예: '13반 - 11차시' → '113'"""
    for name, code in _CLASS_NAME_TO_CODE.items():
        if name in title:
            return code
    return None

def _expected_student_id_by_code(class_code, student_num):
    """반 코드 + 번호로 올바른 학번 계산. 예: '113', '29' → '11329'"""
    try:
        if not class_code:
            return None
        num_int = int(str(student_num).strip())
        expected = f'{class_code}{num_int:02d}'
        return expected if expected in _VALID_IDS else None
    except Exception:
        return None

# 수업명 → Lesson 객체 빠른 조회 딕셔너리
_TITLE_TO_LESSON = {l['title']: l for l in LESSONS}

# (하위 호환) lesson_id 기반 반 이름 반환
def _class_from_lesson_id(lesson_id):
    try:
        lid = int(lesson_id)
        prefix = lid // 100
        mapping = {1: '8반', 2: '9반', 3: '10반', 4: '11반', 5: '12반', 6: '13반'}
        return mapping.get(prefix, f'기타({lid})')
    except Exception:
        return '알 수 없음'


def teacher_report(request):
    """교사용: 학생 피드백 데이터 전체 리포트 (웹)"""
    if not request.session.get('teacher_auth'):
        return redirect('teacher_dashboard')

    error = None
    students_by_class = OrderedDict()

    try:
        sheet = get_sheet()
        rows = sheet.get_all_values()
        headers = rows[0] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []

        # 학생별로 집계
        # 실제 시트: [0]=날짜 [1]=수업명 [2]=학번 [3]=번호 [4]=이름 [5..9]=피드백
        student_map = {}  # student_id → dict
        for row in data_rows:
            if len(row) < 3:
                continue
            date      = row[0] if len(row) > 0 else ''
            title     = row[1] if len(row) > 1 else ''
            s_id      = row[2] if len(row) > 2 else ''
            s_num     = row[3] if len(row) > 3 else ''
            s_name    = row[4] if len(row) > 4 else ''
            summary   = row[5] if len(row) > 5 else ''
            problem   = row[6] if len(row) > 6 else ''
            career    = row[7] if len(row) > 7 else ''
            deeplearn = row[8] if len(row) > 8 else ''
            peer      = row[9] if len(row) > 9 else ''

            if not s_id:
                continue

            class_name = _class_from_title(title)

            if s_id not in student_map:
                student_map[s_id] = {
                    'student_id':   s_id,
                    'student_num':  s_num,
                    'student_name': s_name,
                    'class_name':   class_name,
                    'count':        0,
                    'submissions':  [],
                }
            student_map[s_id]['count'] += 1
            student_map[s_id]['submissions'].append({
                'date':      date,
                'title':     title,
                'summary':   summary,
                'problem':   problem,
                'career':    career,
                'deeplearn': deeplearn,
                'peer':      peer,
            })

        # 반별로 그룹화 → 번호순 정렬
        for s in sorted(student_map.values(),
                        key=lambda x: (x['class_name'], x['student_num'].zfill(3))):
            cn = s['class_name']
            if cn not in students_by_class:
                students_by_class[cn] = []
            students_by_class[cn].append(s)

    except Exception as e:
        error = f'데이터 불러오기 오류: {e}'

    filter_class = request.GET.get('class', '')
    if filter_class and filter_class in students_by_class:
        students_by_class = {filter_class: students_by_class[filter_class]}

    return render(request, 'feedback/teacher_report.html', {
        'students_by_class': students_by_class,
        'error':             error,
        'filter_class':      filter_class,
        'all_classes':       list(OrderedDict(
            (s['class_name'], None)
            for s in student_map.values()
        ).keys()) if 'student_map' in dir() else [],
    })


def ping(request):
    """콜드스타트 방지용 keep-alive 엔드포인트"""
    from django.http import JsonResponse
    return JsonResponse({'status': 'ok', 'time': datetime.now().isoformat()})


def teacher_report_excel(request):
    """교사용: 학생 피드백 데이터 엑셀 다운로드"""
    if not request.session.get('teacher_auth'):
        return redirect('teacher_dashboard')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        sheet_data = get_sheet()
        rows = sheet_data.get_all_values()
        data_rows = rows[1:] if len(rows) > 1 else []

        # 학생별 집계
        # 실제 시트: [2]=학번 [3]=번호 [4]=이름 [5..9]=피드백
        student_map = {}
        for row in data_rows:
            if len(row) < 3:
                continue
            s_id   = row[2] if len(row) > 2 else ''
            s_num  = row[3] if len(row) > 3 else ''
            s_name = row[4] if len(row) > 4 else ''
            title  = row[1] if len(row) > 1 else ''
            if not s_id:
                continue
            class_name = _class_from_title(title)
            if s_id not in student_map:
                student_map[s_id] = {
                    'class': class_name, 'num': s_num, 'name': s_name,
                    'count': 0, 'submissions': []
                }
            student_map[s_id]['count'] += 1
            student_map[s_id]['submissions'].append({
                'date':  row[0], 'title': row[1],
                'summary':   row[5] if len(row) > 5 else '',
                'problem':   row[6] if len(row) > 6 else '',
                'career':    row[7] if len(row) > 7 else '',
                'deeplearn': row[8] if len(row) > 8 else '',
                'peer':      row[9] if len(row) > 9 else '',
            })

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # 기본 시트 삭제

        # ── 스타일 정의 ──
        hdr_fill   = PatternFill('solid', fgColor='2F5496')
        hdr_font   = Font(color='FFFFFF', bold=True, size=11)
        sub_fill   = PatternFill('solid', fgColor='D6E4F7')
        sub_font   = Font(bold=True, size=10)
        wrap_align = Alignment(wrap_text=True, vertical='top')
        center     = Alignment(horizontal='center', vertical='center')
        thin       = Side(style='thin', color='CCCCCC')
        border     = Border(left=thin, right=thin, top=thin, bottom=thin)

        def set_hdr(cell, value):
            cell.value = value
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = center
            cell.border = border

        def set_cell(cell, value, align=None):
            cell.value = value
            cell.alignment = align or wrap_align
            cell.border = border

        # ── 시트 1: 전체 요약 ──
        ws_sum = wb.create_sheet('전체요약')
        cols = ['반', '번호', '학번', '이름', '제출횟수', '제출차시목록']
        for ci, col in enumerate(cols, 1):
            set_hdr(ws_sum.cell(1, ci), col)
        ws_sum.row_dimensions[1].height = 22

        ri = 2
        for s in sorted(student_map.values(),
                        key=lambda x: (x['class'], x['num'].zfill(3))):
            titles = [sub['title'] for sub in s['submissions']]
            row_data = [
                s['class'], s['num'], list(student_map.keys())[list(student_map.values()).index(s)],
                s['name'], s['count'], ', '.join(titles)
            ]
            # student_id 찾기
            sid = next(k for k, v in student_map.items() if v is s)
            row_data[2] = sid
            for ci, val in enumerate(row_data, 1):
                set_cell(ws_sum.cell(ri, ci), val,
                         center if ci <= 5 else wrap_align)
            ri += 1

        ws_sum.column_dimensions['A'].width = 8
        ws_sum.column_dimensions['B'].width = 7
        ws_sum.column_dimensions['C'].width = 12
        ws_sum.column_dimensions['D'].width = 10
        ws_sum.column_dimensions['E'].width = 9
        ws_sum.column_dimensions['F'].width = 50

        # ── 반별 상세 시트 ──
        classes_order = ['8반', '9반', '10반', '11반', '12반', '13반']
        class_students = {cn: [] for cn in classes_order}
        for sid, s in student_map.items():
            cn = s['class']
            if cn in class_students:
                class_students[cn].append((sid, s))

        detail_cols = ['번호', '학번', '이름', '날짜', '차시', '핵심개념3가지', '오류/어려움과해결', '진로연결', '심화학습의지', '칭찬한마디']

        for cn in classes_order:
            students_in_class = sorted(class_students.get(cn, []),
                                       key=lambda x: x[1]['num'].zfill(3))
            if not students_in_class:
                continue

            ws = wb.create_sheet(cn)
            for ci, col in enumerate(detail_cols, 1):
                set_hdr(ws.cell(1, ci), col)
            ws.row_dimensions[1].height = 22

            ri = 2
            for sid, s in students_in_class:
                first = True
                for sub in s['submissions']:
                    num_cell  = ws.cell(ri, 1)
                    id_cell   = ws.cell(ri, 2)
                    name_cell = ws.cell(ri, 3)
                    if first:
                        set_cell(num_cell,  s['num'],  center)
                        set_cell(id_cell,   sid,       center)
                        set_cell(name_cell, s['name'], center)
                        first = False
                    else:
                        set_cell(num_cell,  '', center)
                        set_cell(id_cell,   '', center)
                        set_cell(name_cell, '', center)

                    set_cell(ws.cell(ri, 4),  sub['date'],      center)
                    set_cell(ws.cell(ri, 5),  sub['title'])
                    set_cell(ws.cell(ri, 6),  sub['summary'])
                    set_cell(ws.cell(ri, 7),  sub['problem'])
                    set_cell(ws.cell(ri, 8),  sub['career'])
                    set_cell(ws.cell(ri, 9),  sub['deeplearn'])
                    set_cell(ws.cell(ri, 10), sub['peer'])
                    ws.row_dimensions[ri].height = 60
                    ri += 1

            # 열 너비
            widths = [7, 12, 10, 18, 20, 35, 35, 35, 35, 25]
            for ci, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(ci)].width = w

        # ── HttpResponse 반환 ──
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        today = datetime.now().strftime('%Y%m%d')
        fname = f'feedback_report_{today}.xlsx'
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{fname}"'
        return response

    except ImportError:
        return HttpResponse('openpyxl 패키지가 필요합니다. requirements.txt 확인 후 재배포하세요.', status=500)
    except Exception as e:
        return HttpResponse(f'오류: {e}', status=500)


# ═══════════════════════════════════════════════════════════════
#  학번 검증 및 정정 시스템  (Student ID Validation)
# ═══════════════════════════════════════════════════════════════

# CSV에서 추출한 실제 유효 학번 집합
_VALID_IDS = set(
    f'{cls}{num:02d}'
    for cls, count in [('108', 30), ('109', 30), ('110', 30),
                       ('111', 30), ('112', 29), ('113', 29)]
    for num in range(1, count + 1)
)

# lesson_id 앞자리 → 반 코드 매핑
_LESSON_CLASS_CODE = {1: '108', 2: '109', 3: '110', 4: '111', 5: '112', 6: '113'}
_CLASS_CODE_NAME   = {'108': '8반', '109': '9반', '110': '10반',
                      '111': '11반', '112': '12반', '113': '13반'}


def _expected_student_id(lesson_id, student_num):
    """lesson_id + 번호로 올바른 학번 계산.  예) lesson_id=103, num='15' → '10815'"""
    try:
        prefix = int(lesson_id) // 100
        code   = _LESSON_CLASS_CODE.get(prefix)
        if not code:
            return None
        num_int = int(str(student_num).strip())
        expected = f'{code}{num_int:02d}'
        return expected if expected in _VALID_IDS else None
    except Exception:
        return None


def teacher_validate_ids(request):
    """교사용: 구글 시트 학번 검증 및 정정 페이지"""
    if not request.session.get('teacher_auth'):
        return redirect('teacher_dashboard')

    error = success = None
    rows_info = []   # 분석 결과
    total = ok = wrong = unknown = 0

    try:
        sheet   = get_sheet()
        all_rows = sheet.get_all_values()
        data_rows = all_rows[1:] if len(all_rows) > 1 else []

        for i, row in enumerate(data_rows, start=2):   # 시트 행 번호 (1=헤더)
            if len(row) < 3:
                continue
            date         = row[0]
            title        = row[1]
            entered_id   = row[2].strip()   # C열: 학번
            student_num  = row[3] if len(row) > 3 else ''  # D열: 번호
            student_name = row[4] if len(row) > 4 else ''  # E열: 이름

            if not entered_id:
                continue
            total += 1
            class_code = _class_code_from_title(title)
            expected   = _expected_student_id_by_code(class_code, student_num)

            if expected is None:
                status = 'unknown'
                unknown += 1
            elif entered_id == expected:
                status = 'ok'
                ok += 1
            else:
                status = 'wrong'
                wrong += 1

            rows_info.append({
                'row':          i,
                'date':         date,
                'title':        title,
                'student_num':  student_num,
                'student_name': student_name,
                'entered_id':   entered_id,
                'expected_id':  expected or '알 수 없음',
                'status':       status,
            })

    except Exception as e:
        error = f'구글 시트 읽기 오류: {e}'

    # "정정 적용" POST 처리
    if request.method == 'POST' and not error:
        action = request.POST.get('action')
        if action == 'fix_all':
            try:
                sheet    = get_sheet()
                all_rows = sheet.get_all_values()
                data_rows = all_rows[1:]
                fixed = 0
                for i, row in enumerate(data_rows, start=2):
                    if len(row) < 3:
                        continue
                    title       = row[1]
                    entered_id  = row[2].strip()  # C열: 학번
                    student_num = row[3] if len(row) > 3 else ''
                    class_code  = _class_code_from_title(title)
                    expected    = _expected_student_id_by_code(class_code, student_num)
                    if expected and entered_id != expected:
                        sheet.update_cell(i, 3, expected)   # C열(gspread 1-based) = 학번
                        fixed += 1
                success = f'✓ {fixed}건의 학번이 자동 정정되었습니다.'
                # 목록 갱신
                return redirect('/teacher/validate/?fixed=' + str(fixed))
            except Exception as e:
                error = f'정정 중 오류: {e}'

    fixed_count = request.GET.get('fixed')
    if fixed_count:
        success = f'✓ {fixed_count}건의 학번이 자동 정정되었습니다.'

    wrong_rows   = [r for r in rows_info if r['status'] == 'wrong']
    unknown_rows = [r for r in rows_info if r['status'] == 'unknown']

    return render(request, 'feedback/teacher_validate.html', {
        'rows_info':    rows_info,
        'wrong_rows':   wrong_rows,
        'unknown_rows': unknown_rows,
        'total':   total,
        'ok':      ok,
        'wrong':   wrong,
        'unknown': unknown,
        'error':   error,
        'success': success,
    })


# ═══════════════════════════════════════════════════════════════
#  중복 피드백 제거  (Deduplication)
# ═══════════════════════════════════════════════════════════════

def teacher_dedup(request):
    """교사용: 구글 시트 main 시트에서 중복 피드백 제거 (같은 lesson_id + student_id 중 마지막만 유지)"""
    if not request.session.get('teacher_auth'):
        return redirect('teacher_dashboard')

    error = success = None
    preview_groups = []   # 중복 그룹 미리보기용
    total_dups = 0
    analyzed = False

    try:
        sheet    = get_sheet()
        all_rows = sheet.get_all_values()
        header   = all_rows[0] if all_rows else []
        data_rows = all_rows[1:] if len(all_rows) > 1 else []

        # (수업명, 학번) → [(sheet_row_index, row_data), ...]
        from collections import defaultdict
        key_map = defaultdict(list)
        for i, row in enumerate(data_rows, start=2):  # 시트 행 번호 (헤더=1)
            if len(row) < 3:
                continue
            title      = row[1].strip()   # B열: 수업명
            student_id = row[2].strip()   # C열: 학번
            if not title or not student_id:
                continue
            key_map[(title, student_id)].append((i, row))

        # 중복 그룹만 추출
        dup_groups = {k: v for k, v in key_map.items() if len(v) > 1}
        total_dups = sum(len(v) - 1 for v in dup_groups.values())  # 삭제될 행 수
        analyzed = True

        # 미리보기용 데이터 (최대 20개 그룹)
        for (title, sid), rows_list in list(dup_groups.items())[:20]:
            keep = rows_list[-1]
            delete_rows = rows_list[:-1]
            preview_groups.append({
                'lesson_id':     title,           # 수업명으로 표시
                'student_id':    sid,
                'student_name':  keep[1][4] if len(keep[1]) > 4 else '',  # E열: 이름
                'lesson_title':  title,
                'count':         len(rows_list),
                'keep_row':      keep[0],
                'delete_count':  len(delete_rows),
            })

    except Exception as e:
        error = f'구글 시트 읽기 오류: {e}'

    # POST: 실제 삭제 실행
    if request.method == 'POST' and not error:
        action = request.POST.get('action')
        if action == 'dedup_all':
            try:
                sheet    = get_sheet()
                all_rows = sheet.get_all_values()
                data_rows = all_rows[1:] if len(all_rows) > 1 else []

                from collections import defaultdict
                key_map = defaultdict(list)
                for i, row in enumerate(data_rows, start=2):
                    if len(row) < 3:
                        continue
                    title      = row[1].strip()  # B열: 수업명
                    student_id = row[2].strip()  # C열: 학번
                    if not title or not student_id:
                        continue
                    key_map[(title, student_id)].append(i)  # 시트 행 번호만 저장

                # 삭제할 행 번호 수집 (각 그룹에서 마지막 제외하고 모두)
                rows_to_delete = []
                for v in key_map.values():
                    if len(v) > 1:
                        rows_to_delete.extend(v[:-1])  # 마지막 행만 보존

                # 행 번호 내림차순으로 정렬하여 삭제 (위에서 삭제하면 인덱스 밀림 방지)
                rows_to_delete.sort(reverse=True)
                deleted = len(rows_to_delete)

                for row_num in rows_to_delete:
                    sheet.delete_rows(row_num)

                success = f'✓ {deleted}개의 중복 행이 삭제되었습니다. (각 학생별 마지막 제출 기록 보존)'
                # 분석 데이터 초기화
                preview_groups = []
                total_dups = 0

            except Exception as e:
                error = f'삭제 중 오류: {e}'

    return render(request, 'feedback/teacher_dedup.html', {
        'error':          error,
        'success':        success,
        'preview_groups': preview_groups,
        'total_dups':     total_dups,
        'analyzed':       analyzed,
        'group_count':    len(preview_groups),
    })


def teacher_validate_excel(request):
    """교사용: 학번 검증 결과 엑셀 다운로드"""
    if not request.session.get('teacher_auth'):
        return redirect('teacher_dashboard')
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        sheet    = get_sheet()
        all_rows = sheet.get_all_values()
        data_rows = all_rows[1:] if len(all_rows) > 1 else []

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '학번정정목록'

        thin = Side(style='thin', color='CCCCCC')
        brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
        ctr  = Alignment(horizontal='center', vertical='center')
        wrap = Alignment(wrap_text=True, vertical='center')

        headers = ['시트행', '날짜', '차시', '번호', '이름', '입력학번', '정확한학번', '상태']
        fill_hdr = PatternFill('solid', fgColor='2F5496')
        for ci, h in enumerate(headers, 1):
            c = ws.cell(1, ci, h)
            c.font = Font(color='FFFFFF', bold=True)
            c.fill = fill_hdr
            c.alignment = ctr
            c.border = brd

        fill_ok  = PatternFill('solid', fgColor='D6F5D6')
        fill_bad = PatternFill('solid', fgColor='FFD6D6')
        fill_unk = PatternFill('solid', fgColor='FFF5D6')

        ri = 2
        for row in data_rows:
            if len(row) < 3:
                continue
            title        = row[1]
            entered_id   = row[2].strip()  # C열: 학번
            student_num  = row[3] if len(row) > 3 else ''
            student_name = row[4] if len(row) > 4 else ''
            class_code   = _class_code_from_title(title)
            expected     = _expected_student_id_by_code(class_code, student_num)

            if not entered_id:
                continue
            if expected is None:
                status, fill = '번호오류', fill_unk
            elif entered_id == expected:
                status, fill = '정상', fill_ok
            else:
                status, fill = '오류→정정필요', fill_bad

            vals = [ri - 1, row[0], row[1], student_num,
                    student_name,
                    entered_id, expected or '확인불가', status]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(ri, ci, v)
                c.alignment = ctr if ci != 3 else wrap
                c.border = brd
                c.fill = fill
            ri += 1

        for ci, w in enumerate([7, 18, 22, 7, 10, 12, 12, 14], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 22

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        today = datetime.now().strftime('%Y%m%d')
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="id_validation_{today}.xlsx"'
        return response
    except Exception as e:
        return HttpResponse(f'오류: {e}', status=500)